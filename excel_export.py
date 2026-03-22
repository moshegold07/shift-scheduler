"""
Excel Export — styled like the example screenshot
RTL layout, colored cells per employee, Hebrew headers
"""
from datetime import date, timedelta
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from scheduler import SHIFTS, SHIFT_HOURS, MORNING, EVENING, NIGHT, get_day_name

# Color palette for employees (matching screenshot style)
EMPLOYEE_COLORS = [
    "00B0F0",  # light blue
    "FF00FF",  # pink/magenta
    "FFC000",  # orange/gold
    "00B050",  # green
    "FFFF00",  # yellow
    "FF6600",  # dark orange
    "9966FF",  # purple
    "FF9999",  # light pink
]


def assign_colors(employees: List[str]) -> Dict[str, str]:
    """Assign a color to each employee."""
    colors = {}
    for i, emp in enumerate(employees):
        colors[emp] = EMPLOYEE_COLORS[i % len(EMPLOYEE_COLORS)]
    return colors


def _write_schedule_sheet(
    ws,
    schedule: Dict[Tuple[date, str], str],
    employees: List[str],
    start_date: date,
    num_days: int,
    holidays: Dict[date, str],
    colors: Dict[str, str],
):
    """Write a single schedule sheet (shared by both tabs)."""
    from scheduler import SHIFT_COST

    ws.sheet_view.rightToLeft = True

    dates = [start_date + timedelta(days=i) for i in range(num_days)]

    holiday_fill = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
    erev_hag_fill = PatternFill(start_color="FF8888", end_color="FF8888", fill_type="solid")

    header_font = Font(name="Arial", bold=True, size=12)
    cell_font = Font(name="Arial", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    dark_colors = {"FF00FF", "00B050", "4472C4", "9966FF", "FF6600"}

    # --- Row 1: header ---
    ws.cell(row=1, column=1, value="תאריך\nיום")
    ws.cell(row=1, column=1).font = header_font_white
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = center_align
    ws.cell(row=1, column=1).border = thin_border

    for col_idx, d in enumerate(dates, start=2):
        holiday_name = holidays.get(d, "")
        label = f"{d.strftime('%d.%m')}\n{get_day_name(d)}"
        if holiday_name:
            label += f"\n🔴 {holiday_name}"

        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.alignment = center_align
        cell.border = thin_border

        if holiday_name and "ערב" in holiday_name:
            cell.fill = erev_hag_fill
            cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        elif holiday_name:
            cell.fill = holiday_fill
            cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        else:
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
            cell.font = header_font

    # --- Shift rows ---
    shift_labels = {
        MORNING: f"בוקר\n{SHIFT_HOURS[MORNING]}",
        EVENING: f"ערב\n{SHIFT_HOURS[EVENING]}",
        NIGHT: f"לילה\n{SHIFT_HOURS[NIGHT]}",
    }

    for shift_idx, shift in enumerate(SHIFTS):
        row = 2 + shift_idx

        cell = ws.cell(row=row, column=1, value=shift_labels[shift])
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

        for col_idx, d in enumerate(dates, start=2):
            emp = schedule.get((d, shift), "")
            cell = ws.cell(row=row, column=col_idx, value=emp)
            cell.alignment = center_align
            cell.border = thin_border

            if emp and emp in colors:
                cell.fill = PatternFill(
                    start_color=colors[emp], end_color=colors[emp], fill_type="solid"
                )
                if colors[emp] in dark_colors:
                    cell.font = Font(name="Arial", size=11, color="FFFFFF")
                else:
                    cell.font = Font(name="Arial", size=11)
            else:
                cell.font = Font(name="Arial", size=11)

    # --- Column widths ---
    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, len(dates) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13

    # --- Row heights ---
    ws.row_dimensions[1].height = 52
    for row in range(2, 5):
        ws.row_dimensions[row].height = 48

    # --- Legend ---
    legend_row = 6
    lc = ws.cell(row=legend_row, column=1, value="מקרא:")
    lc.font = Font(name="Arial", bold=True, size=11)
    lc.alignment = center_align

    for i, emp in enumerate(employees):
        col = 2 + i
        cell = ws.cell(row=legend_row, column=col, value=emp)
        cell.fill = PatternFill(
            start_color=colors[emp], end_color=colors[emp], fill_type="solid"
        )
        cell.font = Font(name="Arial", size=11, color="FFFFFF" if colors[emp] in dark_colors else "000000")
        cell.alignment = center_align
        cell.border = thin_border

    # --- Summary ---
    summary_row = 8
    sc = ws.cell(row=summary_row, column=1, value="סה״כ נקודות:")
    sc.font = Font(name="Arial", bold=True, size=11)
    sc.alignment = center_align

    for i, emp in enumerate(employees):
        total = 0
        for d_idx in range(num_days):
            d = start_date + timedelta(days=d_idx)
            for shift in SHIFTS:
                if schedule.get((d, shift)) == emp:
                    total += SHIFT_COST[shift]
        col = 2 + i
        cell = ws.cell(row=summary_row, column=col, value=f"{emp}: {total}")
        cell.font = Font(name="Arial", size=11)
        cell.alignment = center_align
        cell.border = thin_border


def export_to_excel(
    schedule_internal: Dict[Tuple[date, str], str],
    schedule_external: Dict[Tuple[date, str], str],
    employees: List[str],
    start_date: date,
    num_days: int,
    output_path: str = "משמרות.xlsx",
    holidays: Optional[Dict[date, str]] = None,
):
    """
    Export two schedules to a single Excel file with two sheets:
      - "לשלישות"  — external/distributed schedule (no double shifts per employee per day)
      - "פנימי"    — internal schedule (double shifts allowed except on Friday)
    """
    if holidays is None:
        holidays = {}

    wb = Workbook()
    colors = assign_colors(employees)

    # Sheet 1: לשלישות (external)
    ws_ext = wb.active
    ws_ext.title = "לשלישות"
    _write_schedule_sheet(ws_ext, schedule_external, employees, start_date, num_days, holidays, colors)

    # Sheet 2: פנימי (internal)
    ws_int = wb.create_sheet(title="פנימי")
    _write_schedule_sheet(ws_int, schedule_internal, employees, start_date, num_days, holidays, colors)

    wb.save(output_path)
    return output_path
